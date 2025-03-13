"""
Script Name: PvSimulation.py
Author: Khalid Alrehaili
Date: [2025-03-14]
Description: This script is a simulation for location assessing
             and the calucation of optimial azimuth and tilt for
             PV module.
"""

from openpyxl import load_workbook
import datetime
from pvlib import solarposition
import pandas as pd
import numpy as np
from numpy import sin, cos, acos, radians, pi
import matplotlib.pyplot as plt
from alive_progress import alive_bar # for progress bar
import time
import pytz
start_time = time.time()

# Location Coordinates
lat = 24.37 # Change it to the latitude of the location
lon = 39.50 # Change it to the longitude of the location
tz = pytz.timezone('Asia/Riyadh') # Change it to the timezone used in the location under interest

# Collecting Metrological Data From ExcelSheet "medina_2019_weather_data.xlsx"
workbook = load_workbook(filename="medina_2019_weather_data.xlsx")
sheet = workbook.active
MetrologicalData = []
for column in sheet.iter_cols(values_only=True, min_row=2):
    MetrologicalData.append(column)

#Time Data
Years = MetrologicalData[0]
Months = MetrologicalData[1]
Days = MetrologicalData[2]
Hours = MetrologicalData[3]
Minuts = MetrologicalData[4]

# Storing each time point in datetime objects

DateList = np.array([]) # list containing datetime object for all time points.
for i in range(0, 8760):
    DateList = np.append(DateList, tz.localize(datetime.datetime(Years[i], Months[i], Days[i], Hours[i], Minuts[i])))

#GHI, DHI, DNI Data
GHItemp = MetrologicalData[6]  # ClearSky GHI
GHItemp = pd.DataFrame(GHItemp, index = DateList)

alpha = 0.2

DHItemp = MetrologicalData[9]  # ClearSky DHI
DHItemp = pd.DataFrame(DHItemp, index = DateList)

DNItemp = MetrologicalData[12] # ClearSky DNI
DNItemp = pd.DataFrame(DNItemp, index = DateList)
print(DHItemp)
#Calculating Sun Position All Year Around each hour
solpos = solarposition.get_solarposition(DateList, lat, lon)
solpos = solpos.loc[solpos['elevation'] > 0, :] # Remove sun positions for negative elevations.
dayList = solpos.index # A subset from DateList that contains only times where the sun is above the horizon!

#optimization process by converting data to np.array objects and removing night time values
GHI = np.array([])
DHI = np.array([])
DNI = np.array([])
sAlti = np.array([])
sAzim = np.array([])
sAltiCos = np.array([])
sAltiSin = np.array([])

for day1 in dayList:
    GHI = np.append(GHI, GHItemp.loc[day1].values)
    DHI = np.append(DHI, DHItemp.loc[day1].values)
    DNI = np.append(DNI, DNItemp.loc[day1].values)

    x = solpos.loc[day1]
    sAlti = np.append(sAlti, x.elevation)
    sAzim = np.append(sAzim, x.azimuth)
    sAltiCos = np.append(sAltiCos ,cos(radians(x.elevation)))
    sAltiSin = np.append(sAltiSin, sin(radians(x.elevation)))
    
# Loop through each altitude angle and direction of module and calculate total irradiance.
angleSteps = 1
ModuleTilt = np.arange(0,90,angleSteps) # from 0 to 90
ModuleAzimuth = np.arange(0,360,angleSteps) # from 0 to 360
answer_matrix = np.zeros((int(90/angleSteps), int(360/angleSteps)), dtype=float)

with alive_bar(len(ModuleTilt)) as bar:
    for mTilt in ModuleTilt:
        bar()
        mAlti = 90 - mTilt # Module Tilt Calculation.
        SVF = 1/2 + 1/2*cos(radians(mTilt)) # Sky View Factor (SVF) Calculation.
        cosmAlti = cos(radians(mAlti))
        sinmAlti = sin(radians(mAlti))

        for mAzim in ModuleAzimuth:
            tempCos = cos(radians(mAzim - sAzim))
            cosAOI = cosmAlti*sAltiCos*tempCos + sinmAlti*sAltiSin
            cosAOI[cosAOI < 0] = 0 # Reset AOI values to 0 when the sun is behind the module
            # calculating Irradiance components with vectors (np.array objects)
            Gdirect = cosAOI * DNI
            Gdiffused = SVF * DHI
            Galbedo = GHI * alpha * (1 - SVF)
            Gtotal = Gdirect + Gdiffused + Galbedo

            # Appending Answer to the final matrix
            sum1 = np.sum(Gtotal) / 1000
            answer_matrix[int(mTilt/angleSteps)][int(mAzim/angleSteps)] = sum1
        

print("RunTime: ")
print("--- %s seconds ---" % round(time.time() - start_time, 2))
tempPostition = answer_matrix.argmax()
OptimalPosition = np.unravel_index(tempPostition, answer_matrix.shape)

print("Optimal Module Tilt:", f"{OptimalPosition[0]*angleSteps}{'\N{DEGREE SIGN}'}")
print("Optimal Module Azimuth:", f"{OptimalPosition[1]*angleSteps}{'\N{DEGREE SIGN}'}")
print("Maximum Calculated Energy:", round(answer_matrix.max(), 2), " kWh")
print("Minimum Calcualted Energy:", round(answer_matrix.min(), 2), " kWh")

x = plt.contourf(answer_matrix, np.linspace(answer_matrix.min(), answer_matrix.max(), 200), cmap="jet")
plt.colorbar(x)
plt.show()
